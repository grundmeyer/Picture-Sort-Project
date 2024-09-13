"""Sort pictures by location A-J and K-Z"""
import os
import shutil
import time
import math
from urllib.parse import unquote
from openpyxl import load_workbook
from PIL import Image
from colorama import Fore
# The following import can be either "config_update" or "config_checklist"
from config_checklist import WORKING_DIRECTORY, SPREADSHEET_PATH, PICTURE_ROOT, SCALE_FACTOR, FIRST_ROW, LAST_ROW, CATEGORY_LIST

    # 1. Two folders: a-j and k-z
    # 2. -> subfolders for each location
    # 3. --> subfolders for wall, rack, etc.
    # 4. Run and upload

def writeImage(link, location, imageFolder, imageType, half, number):
    temp1 = link.split("/")
    photoName = unquote(temp1[len(temp1) - 1])

    if "iowa-my.sharepoint" in link:
        print(f"{Fore.YELLOW}Photos located in a personal OneDrive folder.{Fore.RESET}")
        return

    if len(photoName) < 3:
        return
    
    if half:
        os.makedirs(Rf"{half}\{location}\{imageType}", exist_ok=True)
    else:
        os.makedirs(Rf"{location}\{imageType}", exist_ok=True)
    
    if ".MOV" in photoName:
        shutil.copy(Rf"{PICTURE_ROOT}\{imageFolder}\{photoName}", Rf"{half}\{location}\{imageType}\{imageType}-{number}.MOV")
        return
    
    with Image.open(Rf"{PICTURE_ROOT}\{imageFolder}\{photoName}") as sourceFile:
        width, height = sourceFile.size
        resizedImage = sourceFile.resize(
            (math.floor(width * SCALE_FACTOR), math.floor(height * SCALE_FACTOR)),
            Image.Resampling.LANCZOS
        )
        resizedImage.save(
            Rf"{half}\{location}\{imageType}\{imageType}-{number}.jpg" if half else Rf"{location}\{imageType}\{imageType}-{number}.jpg",
            "JPEG",
            quality=85,
            optimize=True,
            exif=sourceFile.info.get("exif")
        )

def main():
    start = time.time()
    os.chdir(WORKING_DIRECTORY)

    wb = load_workbook(filename=SPREADSHEET_PATH)
    ws = wb.active

    # Iterate by row
    for row in range(FIRST_ROW, LAST_ROW + 1):
        i = 1
        location = f"Unknown Location {i}"
        half = None

        # Get location
        if ws[f"Q{row}"].value is not None:
            if ws[f"G{row}"].value is not None:
                location = (ws[f"G{row}"].value).replace("\t", " ")
                half = "Location A-J"
            elif ws[f"H{row}"].value is not None:
                location = (ws[f"H{row}"].value).replace("\t", " ")
                half = "Location K-Z"
        else:
            continue

        # Iterate by wall, rack, etc.
        for category in CATEGORY_LIST:
            j = 1
            linksList = []
            if ws[f"{category[0]}{row}"].value is not None:
                linksList = (ws[f"{category[0]}{row}"].value).replace(" ", "").split(";")
            # Iterate by link (photo) for wall, rack, etc.
            for link in linksList:
                writeImage(link, location, category[2], category[1], half, j)
                j += 1
        print(f"Photos in row {row} saved.")
        i += 1

    print(f"Completed in {Fore.GREEN}{time.time() - start:.2f}{Fore.RESET} seconds.")

if __name__ == "__main__":
    main()
