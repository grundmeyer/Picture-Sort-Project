from urllib.parse import unquote
from io import BytesIO
from datetime import datetime, timezone, timedelta
import requests, os, math, time
from openpyxl import load_workbook
from PIL import Image
from config import PROD_URL, FIRST_ROW, LAST_ROW, SPREADSHEET_PATH, PICTURE_ROOT, SCALE_FACTOR, HEADERS

def main():
    os.chdir(PICTURE_ROOT)
    start = time.time()

    # Get building list from Netbox
    data = requests.get(url=f"{PROD_URL}api/dcim/locations/", headers=HEADERS, verify=False, timeout=5)
    d = data.json()
    buildingList = {}
    populateDict(d, buildingList)

    while d['next']:
        data = requests.get(url=d['next'], headers=HEADERS, verify=False, timeout=5)
        d = data.json()
        populateDict(d, buildingList)

    # Load spreadsheet
    wb = load_workbook(filename=SPREADSHEET_PATH)
    ws = wb.active

    # Create lists for each set of pictures
    wallLinks = {}
    rackLinks = {}
    powerLinks = {}
    fiberLinks = {}
    switchLinks = {}
    doorLinks = {}

    linksList = [
        (wallLinks,   "Q", "Wall",       "Upload Wall and Celing Pictures"),
        (rackLinks,   "R", "Rack",       "Upload Rack Pictures"),
        (powerLinks,  "T", "Power",      "Upload Power Panel Pictures"),
        (fiberLinks,  "U", "FiberPanel", "Upload Fiber Panel Pictures"),
        (switchLinks, "S", "Switch",     "Upload SWITCH Pictures"),
        (doorLinks,   "K", "Door",       "Upload Door Pictures")
    ]

    count = 1
    total = 0

    for i in range(FIRST_ROW, LAST_ROW + 1):
        if ws[f'Q{i}'].value:
            for group in linksList:
                total += populateLinks(i, ws, group[0], group[1])

    for i in range(FIRST_ROW, LAST_ROW + 1):
        if i in wallLinks:
            tc = ((ws[f"G{i}"].value).replace("\t", " ") if ws[f"G{i}"].value
                  else (ws[f"H{i}"].value).replace("\t", " "))

            closet = ""
            temp3 = tc.split()
            last_index = len(temp3) - 1
            for item in buildingList:
                temp2 = item.split('-')
                if temp2[0] == temp3[last_index] and temp2[1] in tc:
                    closet = item

            if tc == "NAGLE FAMILY CLUB HOUSEEB015NEWFGCH":
                closet = "FGCH-EB01"
            elif tc == "PARKING LOT 49 - MELROSE SURFA EB01  L49":
                closet = "LOT49-EB01"
            elif tc == "English-Philosophy Building":
                closet = "EPB-EB01"

            uploadedImages = {}
            x = requests.get(
                url=f"{PROD_URL}api/extras/image-attachments",
                headers=HEADERS,
                params={"object_id": buildingList[closet]['id']},
                timeout=5,
            )
            y = x.json()
            for z in y['results']:
                uploadedImages[z['display']] = z['id']

            for b in linksList:
                for c in b[0][i]:
                    pushImage(
                        c,
                        closet,
                        buildingList,
                        b[2],
                        b[3],
                        SCALE_FACTOR,
                        PROD_URL,
                        HEADERS,
                        uploadedImages
                    )
                    count = updateProgress(count, total)
 
    print(f"\nCompleted in {time.time() - start:.2f} seconds.")


def populateDict(response, closetList):
    for closet in response['results']:
        temp = {"id": closet['id'], "Wall": 0, "Rack": 0, "Power": 0, "FiberPanel": 0, "Switch": 0, "Door": 0}
        closetList[closet['name']] = temp


def populateLinks(index, worksheet, linkDict, column):
    linkDict[index] = (worksheet[f"{column}{index}"].value).replace(" ", "").split(";")
    return len(linkDict[index])


def isUploaded(existsList, displayName, url, headers, folderName, imageName):
    try:
        id = existsList[displayName]
        test = requests.get(f"{url}api/extras/image-attachments/{id}/", headers=headers, timeout=5)
        if (test.status_code == 404): # Unhandled exception, just skip it
            print(f"{displayName} was not found.")
            return True
        elif (test.status_code == 200): # Image is on Netbox, check date
            testInfo = test.json()

            mTime = os.path.getmtime(f"{folderName}\\{imageName}")

            # find if the file was changed in DST
            localTime = time.localtime(mTime)
            if localTime.tm_isdst:
                localtz = timezone(timedelta(hours=-5))
            else:
                localtz = timezone(timedelta(hours=-6))
            
            fileModifiedTime = datetime.fromtimestamp((mTime), localtz)

            uploadTime = datetime.fromisoformat(testInfo.get("last_updated"))

            if (fileModifiedTime > uploadTime): # The image file is newer than Netbox
                deleteRequest = requests.delete(
                    f"{url}api/extras/image-attachments/{id}/",
                    headers=headers,
                    timeout=5
                )
                if (deleteRequest.status_code == 204):
                    print(f"{displayName} is out of date. Old file deleted.")
                return False
            return True
        
    except KeyError: # The image is not on Netbox
        return False
    

def pushImage(
        link,
        closet,
        closetList,
        imageType,
        folderName,
        scaleFactor,
        url,
        headers,
        existsList):
    closetList[closet][imageType] += 1
    displayName = f"{closet}-{imageType}_{closetList[closet][imageType] // 10}{closetList[closet][imageType] % 10}.jpg"
    temp1 = link.split("/")
    imageName = unquote(temp1[len(temp1) - 1])

    # the Netbox image is up to date
    if isUploaded(existsList, displayName, url, headers, folderName, imageName):
        return

    # the image is not on Netbox or outdated
    
    with open(f"{folderName}\\{imageName}", "rb") as sourceImage:
        print(f"{displayName} is not uploaded or out of date.")

        # Load image to buffer
        sourceBytes = BytesIO(sourceImage.read())

        # Open with pillow
        originalImage = Image.open(sourceBytes)
        width, height = originalImage.size

        # Resize with pillow
        compressedImage = originalImage.resize(
            (math.floor(width * scaleFactor), math.floor(height * scaleFactor)),
            Image.Resampling.LANCZOS
        )
        width, height = compressedImage.size

        # Write compressed image to new buffer
        compressedBuffer = BytesIO()
        compressedImage.save(
            compressedBuffer,
            "JPEG",
            quality=85,
            optimize=True,
            exif=originalImage.info.get("exif")
        )

        # Upload image in buffer to Netbox
        files = {
            "image": (displayName, compressedBuffer.getvalue())
        }

        body = {
            "content_type": "dcim.location",
            "object_id": closetList[closet]["id"],
            "image_height": height,
            "image_width": width
        }

        response = requests.post(
            f"{url}api/extras/image-attachments/",
            headers=headers,
            files=files,
            data=body,
            timeout=10
        )

        if response.status_code == 201:
            print(f"{displayName} done.")
        else:
            print(f"Error sending {displayName}.")


def updateProgress(count, total):
    print(f"{count} of {total} images uploaded.\r", end='')
    return count + 1


if __name__ == "__main__":
    main()
